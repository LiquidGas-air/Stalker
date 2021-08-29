import openpyxl
friends= {}
PATH = r"E:\Programms2\gathering\data.xlsx"

excel = openpyxl.load_workbook(PATH)
sheet = excel.active
max_row = sheet.max_row
for i in range(1, max_row+1):
    friends[sheet.cell(column = 1, row = i ).value] ={}
for i in range(1, max_row+1):
    friends[ sheet.cell(column = 1, row = i ).value ].update( {sheet.cell(column = 2, row = i).value : sheet.cell(column = 3, row = i).value})
    print(i)
print(friends)